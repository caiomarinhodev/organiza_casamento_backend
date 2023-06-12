import os

import openpyxl
from django.http import HttpResponse
from django.template.loader import get_template
from rest_framework.decorators import api_view
from xhtml2pdf import pisa

from app.models import Event, Guest


@api_view(['GET'])
def event_report(request, event_id):
    # Busca o evento pelo ID
    evento = Event.objects.get(id=event_id)

    # Cria um novo arquivo Excel
    wb = openpyxl.Workbook()

    # Seleciona a primeira planilha
    ws = wb.active

    # Define o cabeçalho
    ws['A1'] = 'Nome do Evento'
    ws['B1'] = 'Data'
    ws['C1'] = 'Tamanho'
    ws['D1'] = 'Estilo'
    ws['E1'] = 'Orçamento'
    ws['F1'] = 'Número de Convidados'

    # Preenche os dados do evento na planilha
    ws['A2'] = evento.name
    ws['B2'] = evento.date.strftime('%d/%m/%Y')
    ws['C2'] = evento.size
    ws['D2'] = evento.style
    ws['E2'] = str(evento.budget)
    ws['F2'] = str(evento.guests)

    # Define o nome do arquivo de saída
    filename = f'{evento.name}_relatorio.xlsx'

    # Define o cabeçalho da resposta HTTP
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Salva o arquivo Excel na resposta HTTP
    wb.save(response)

    return response


@api_view(['GET'])
def guest_list(request, event_id):
    # Busca os convidados do evento pelo id
    event = Event.objects.get(id=event_id)
    guests = Guest.objects.filter(event_id=event_id).order_by('-confirmed')

    # Define o nome do arquivo PDF
    filename = f"{event.name} - Lista de Convidados.pdf"

    # Carrega o template HTML para a lista de convidados
    template = get_template('guest_list.html')

    # Define o contexto para o template HTML
    context = {'guests': guests}

    # Renderiza o template HTML com o contexto
    html = template.render(context)

    # Cria um novo arquivo PDF
    pdf = open(filename, 'wb')

    # Cria o PDF a partir do HTML
    pisa.CreatePDF(html, dest=pdf)

    # Fecha o arquivo PDF
    pdf.close()

    # Lê o arquivo PDF como um objeto de bytes
    with open(filename, 'rb') as f:
        pdf_data = f.read()

    # Remove o arquivo PDF
    os.remove(filename)

    # Cria uma resposta HTTP com o arquivo PDF
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
